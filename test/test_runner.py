import xml.etree.ElementTree as ET
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook

# Import the modules
from quickbook_connector.model import Customer, Conflict, ComparisonReport
from quickbook_connector import qb_gateway
from quickbook_connector.excel_reader import extract_customers


# --------------------------------------------------------------------
# MODEL TESTS
# --------------------------------------------------------------------
def test_customer_str():
    c = Customer(record_id="101", name="Acme Corp", source="excel")
    assert "Acme Corp" in str(c)
    assert "101" in str(c)


def test_conflict_fields():
    conf = Conflict(record_id="1", excel_name="A", qb_name="B", reason="name_mismatch")
    assert conf.reason == "name_mismatch"
    assert conf.qb_name == "B"


def test_comparison_report_defaults():
    report = ComparisonReport()
    assert report.excel_only == []
    assert report.conflicts == []


# --------------------------------------------------------------------
# EXCEL READER TESTS
# --------------------------------------------------------------------
def test_extract_customers_valid(tmp_path):
    """Ensure customers are extracted properly from a valid Excel file."""
    # Create a temporary Excel file
    workbook_path = tmp_path / "customers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "customers"
    ws.append(["ID", "Name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    wb.save(workbook_path)

    customers = extract_customers(workbook_path)
    assert len(customers) == 2
    assert customers[0].name == "Alice"
    assert customers[1].record_id == "2"


def test_extract_customers_missing_file():
    """Expect FileNotFoundError for invalid path."""
    with pytest.raises(FileNotFoundError):
        extract_customers(Path("nonexistent.xlsx"))


def test_extract_customers_missing_sheet(tmp_path):
    """Expect ValueError when 'customers' worksheet not found."""
    workbook_path = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.create_sheet("wrong_sheet")
    wb.save(workbook_path)

    with pytest.raises(ValueError):
        extract_customers(workbook_path)


# --------------------------------------------------------------------
# QB CONNECTOR TESTS (Mocked)
# --------------------------------------------------------------------
@pytest.fixture
def fake_qbxml_response():
    """Return a fake CustomerQuery response XML."""
    return """<?xml version="1.0"?>
    <QBXML>
      <QBXMLMsgsRs>
        <CustomerQueryRs statusCode="0" statusMessage="OK" statusSeverity="Info">
          <CustomerRet>
            <Fax>123</Fax>
            <FullName>Test Customer</FullName>
          </CustomerRet>
        </CustomerQueryRs>
      </QBXMLMsgsRs>
    </QBXML>"""


def test_parse_response_success(fake_qbxml_response):
    root = qb_gateway._parse_response(fake_qbxml_response)
    assert root.tag == "QBXML"


def test_parse_response_error():
    bad_xml = """<?xml version="1.0"?>
    <QBXML><QBXMLMsgsRs>
    <CustomerQueryRs statusCode="500" statusMessage="Server Error"/>
    </QBXMLMsgsRs></QBXML>"""
    with pytest.raises(RuntimeError):
        qb_gateway._parse_response(bad_xml)


@patch("quickbook_connector.qb_gateway._send_qbxml")
def test_fetch_customers(mock_send, fake_qbxml_response):
    """Mock QuickBooks call to test customer fetch."""
    mock_send.return_value = ET.fromstring(fake_qbxml_response)

    customers = qb_gateway.fetch_customers()
    assert len(customers) == 1
    assert customers[0].name == "Test Customer"
    assert customers[0].record_id == "123"
    assert customers[0].source == "quickbooks"


@patch("quickbook_connector.qb_gateway._send_qbxml")
def test_add_customer_success(mock_send):
    """Ensure add_customer returns valid Customer object on success."""
    mock_response = """<?xml version="1.0"?>
    <QBXML><QBXMLMsgsRs>
    <StandardTermsAddRs statusCode="0" statusMessage="OK">
      <StandardTermsRet>
        <StdDiscountDays>10</StdDiscountDays>
        <Name>Net 10</Name>
      </StandardTermsRet>
    </StandardTermsAddRs></QBXMLMsgsRs></QBXML>"""
    mock_send.return_value = ET.fromstring(mock_response)

    c = Customer(record_id="10", name="Net 10", source="quickbooks")
    result = qb_gateway.add_customer(None, c)
    assert result.name == "Net 10"
    assert result.record_id == "10"


@patch("quickbook_connector.qb_gateway._send_qbxml")
def test_add_customer_batch_success(mock_send):
    """Ensure batch creation returns valid Customer list."""
    mock_response = """<?xml version="1.0"?>
    <QBXML><QBXMLMsgsRs>
      <StandardTermsAddRs statusCode="0">
        <StandardTermsRet>
          <StdDiscountDays>30</StdDiscountDays>
          <Name>Net 30</Name>
        </StandardTermsRet>
      </StandardTermsAddRs>
    </QBXMLMsgsRs></QBXML>"""
    mock_send.return_value = ET.fromstring(mock_response)

    customers = [Customer(record_id="30", name="Net 30", source="quickbooks")]
    results = qb_gateway.add_customer_batch(None, customers)
    assert len(results) == 1
    assert results[0].name == "Net 30"


def test_escape_xml():
    """Verify that XML escaping works correctly."""
    s = "A&B<C>'\""
    escaped = qb_gateway._escape_xml(s)
    assert "&amp;" in escaped
    assert "&lt;" in escaped
    assert "&apos;" in escaped

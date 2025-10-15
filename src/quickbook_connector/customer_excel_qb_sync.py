"""Excel processing module for reading customer data and QuickBooks integration.

This module provides functions to read customers from Excel,
compare them with QuickBooks Desktop via COM API, and synchronize them.
"""

from dataclasses import dataclass
from typing import Any

import win32com.client
from openpyxl import load_workbook
import xml.etree.ElementTree as ET


@dataclass
class Customer:
    """Represents a customer with Name, Term, and ID."""
    name: str
    term: str
    customer_id: int


@dataclass
class CustomerComparison:
    """Results of comparing Excel and QuickBooks customers."""
    same_id_diff_data: list[tuple[str, str, str, int]]  # (excel_name, qb_name, excel_term, id)
    only_in_excel: list[Customer]  # Customers to add to QB
    only_in_qb: list[Customer]  # Customers in QB but not Excel
    matching_count: int  # Same ID & same data


#  Excel Reading
def read_customers_from_excel(file_path: str) -> list[Customer]:
    """Read customers from Excel file.

    Expected format:
    Sheet name: 'customers'
    Columns:
      - A: Name
      - B: Term
      - C: ID
    """
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook["customers"]
    customers = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, term, cid = row[0], row[1], row[2]
        if not name or not cid:
            continue

        try:
            cid_int = int(cid)
            term_str = str(term).strip() if term else ""
            customers.append(Customer(name=str(name).strip(), term=term_str, customer_id=cid_int))
        except (ValueError, TypeError):
            continue

    return customers


#  QuickBooks Connection
def connect_to_quickbooks() -> Any:
    """Connect to QuickBooks Desktop."""
    try:
        qb_app = win32com.client.Dispatch("QBXMLRP2.RequestProcessor")
        qb_app.OpenConnection("", "Customer Import")
        session = qb_app.BeginSession("", 2)
        return qb_app, session
    except Exception as e:
        print(f"QuickBooks connection error: {str(e)}")
        raise

#  QuickBooks Read / Write
def get_qb_customers() -> list[Customer]:
    """Read customers from QuickBooks."""
    qb_app = None
    session = None
    customers = []

    try:
        qb_app, session = connect_to_quickbooks()
        qbxml_query = """<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="13.0"?>
<QBXML>
    <QBXMLMsgsRq onError="continueOnError">
        <CustomerQueryRq>
            <IncludeRetElement>Name</IncludeRetElement>
            <IncludeRetElement>TermsRef</IncludeRetElement>
            <IncludeRetElement>Fax</IncludeRetElement>
        </CustomerQueryRq>
    </QBXMLMsgsRq>
</QBXML>"""

        response = qb_app.ProcessRequest(session, qbxml_query)
        root = ET.fromstring(response)

        for cust_ret in root.findall(".//CustomerRet"):
            name_elem = cust_ret.find("Name")
            fax_elem = cust_ret.find("Fax")
            term_ref = cust_ret.find("TermsRef/FullName")

            if name_elem is not None and fax_elem is not None:
                try:
                    cid = int(fax_elem.text)
                    term_value = term_ref.text.strip() if term_ref is not None else ""
                    customers.append(
                        Customer(
                            name=name_elem.text.strip(),
                            term=term_value,
                            customer_id=cid,
                        )
                    )
                except (ValueError, TypeError):
                    continue

        return customers

    except Exception as e:
        raise RuntimeError(f"Failed to read QuickBooks customers: {str(e)}") from e
    finally:
        if qb_app and session:
            qb_app.EndSession(session)
            qb_app.CloseConnection()


def create_customers_batch_qbxml(customers: list[Customer]) -> str:
    """Create QBXML for adding multiple customers."""
    requests = []
    for cust in customers:
        name = cust.name.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        term = cust.term.replace("&", "&amp;") if cust.term else ""
        req = f"""        <CustomerAddRq>
            <CustomerAdd>
                <Name>{name}</Name>
                <Fax>{cust.customer_id}</Fax>
                {'<TermsRef><FullName>' + term + '</FullName></TermsRef>' if term else ''}
            </CustomerAdd>
        </CustomerAddRq>"""
        requests.append(req)

    return f"""<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="13.0"?>
<QBXML>
    <QBXMLMsgsRq onError="continueOnError">
{chr(10).join(requests)}
    </QBXMLMsgsRq>
</QBXML>"""


def save_customers_to_quickbooks(customers: list[Customer]) -> list[str]:
    """Add customers to QuickBooks."""
    if not customers:
        return []

    qb_app, session = connect_to_quickbooks()
    try:
        qbxml = create_customers_batch_qbxml(customers)
        response = qb_app.ProcessRequest(session, qbxml)
        root = ET.fromstring(response)

        created = []
        for add_rs in root.findall(".//CustomerAddRs"):
            code = add_rs.get("statusCode")
            if code == "0":
                name_elem = add_rs.find(".//Name")
                if name_elem is not None:
                    created.append(name_elem.text)
            elif code == "3100":
                # Already exists
                pass
            else:
                msg = add_rs.get("statusMessage", "Unknown error")
                print(f"Warning: Failed to create customer: {msg}")

        return created
    finally:
        qb_app.EndSession(session)
        qb_app.CloseConnection()

#  Comparison Logic
def compare_customers(excel_customers: list[Customer], qb_customers: list[Customer]) -> CustomerComparison:
    """Compare customers from Excel and QB."""
    excel_map = {c.customer_id: c for c in excel_customers}
    qb_map = {c.customer_id: c for c in qb_customers}

    same_id_diff_data = []
    only_in_excel = []
    only_in_qb = []
    matching_count = 0

    # Compare IDs
    for cid, excel_cust in excel_map.items():
        if cid in qb_map:
            qb_cust = qb_map[cid]
            if (excel_cust.name.strip().lower() == qb_cust.name.strip().lower() and
                excel_cust.term.strip().lower() == qb_cust.term.strip().lower()):
                matching_count += 1
            else:
                same_id_diff_data.append(
                    (excel_cust.name, qb_cust.name, excel_cust.term, cid)
                )
        else:
            only_in_excel.append(excel_cust)

    # Customers in QB but not Excel
    for cid, qb_cust in qb_map.items():
        if cid not in excel_map:
            only_in_qb.append(qb_cust)

    return CustomerComparison(
        same_id_diff_data=same_id_diff_data,
        only_in_excel=only_in_excel,
        only_in_qb=only_in_qb,
        matching_count=matching_count,
    )


#  Orchestration
def process_customers(file_path: str) -> CustomerComparison:
    """Read customers from Excel, compare with QuickBooks, and synchronize."""
    excel_customers = read_customers_from_excel(file_path)
    if not excel_customers:
        raise ValueError("No customers found in Excel file")

    print(f"Found {len(excel_customers)} customers in Excel")
    print("Reading customers from QuickBooks...")

    qb_customers = get_qb_customers()
    print(f"Found {len(qb_customers)} customers in QuickBooks")

    comparison = compare_customers(excel_customers, qb_customers)

    print("\n=== Comparison Results ===")
    print(f"\nMatching customers (same ID and same data): {comparison.matching_count}")

    if comparison.same_id_diff_data:
        print(f"\nConflicts (same ID, different data): {len(comparison.same_id_diff_data)}")
        for excel_name, qb_name, term, cid in comparison.same_id_diff_data:
            print(f"  ID {cid}: Excel='{excel_name}' ({term}) vs QB='{qb_name}'")
    else:
        print("\nNo conflicts found with same IDs")

    if comparison.only_in_qb:
        print(f"\nCustomers only in QuickBooks (missing in Excel): {len(comparison.only_in_qb)}")
        for c in comparison.only_in_qb:
            print(f"  ID {c.customer_id}: {c.name}")
    else:
        print("\nNo QB-only customers")

    if comparison.only_in_excel:
        print(f"\nAdding {len(comparison.only_in_excel)} new customers to QuickBooks...")
        for c in comparison.only_in_excel:
            print(f"  - {c.name} (ID: {c.customer_id})")

        created = save_customers_to_quickbooks(comparison.only_in_excel)
        print(f"\nSuccessfully created {len(created)} customers in QuickBooks")
    else:
        print("\nNo new customers to add")

    return comparison
